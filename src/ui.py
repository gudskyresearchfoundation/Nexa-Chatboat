import sys, os, io, base64, subprocess
_SRC = os.path.dirname(os.path.abspath(__file__))
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)

import streamlit as st

# ── Auto-install missing packages at runtime (Lightning.ai safe) ──────────────
def _try_install(package: str, import_name: str = None):
    """Attempt pip install if import fails — silent on Lightning.ai."""
    name = import_name or package
    try:
      
        __import__(name)
        return True
    except ImportError:
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", package, "-q"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
            return True
        except Exception:
            return False

# Try to install deps silently
_try_install("pdfplumber")
_try_install("PyPDF2", "PyPDF2")
_try_install("python-docx", "docx")
_try_install("pymupdf", "fitz")

from backend.auth import authenticate, ADMIN_INFO
from model.llm import (
    check_ollama_running, check_model_available,
    get_model_info, ConversationManager, stream_response, MODEL_NAME,
)

# ── Page config — must be first Streamlit call ────────────────────────────────
st.set_page_config(
    page_title="NEXA Chatbot · LLaMA 3",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Avatars ───────────────────────────────────────────────────────────────────
AVATAR = {
    "admin": "https://ui-avatars.com/api/?name=Admin+User&background=6d28d9&color=fff&size=128&bold=true&rounded=true",
    "demo":  "https://ui-avatars.com/api/?name=Demo+User&background=0891b2&color=fff&size=128&bold=true&rounded=true",
}

# ── Session defaults ──────────────────────────────────────────────────────────
_DEFAULTS = {
    "logged_in":       False,
    "username":        None,
    "admin_info":      None,
    "conversation":    None,
    "messages":        [],
    "login_error":     "",
    "page":            "chat",
    "uploaded_file":   None,
    "file_content":    None,
    "file_name":       None,
    "file_type":       None,
}

def _init_state():
    for k, v in _DEFAULTS.items():
        if k not in st.session_state:
            st.session_state[k] = v


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  FILE READING HELPERS  (Lightning.ai safe)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _read_pdf(raw: bytes) -> str:
    """Try every available PDF reader, fall back gracefully."""
    # Method 1 — pdfplumber (best quality)
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
            text = "\n".join(pages).strip()
            if text:
                return text
    except Exception:
        pass

          

    # Method 3 — PyPDF2
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(raw))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n".join(pages).strip()
        if text:
            return text
    except Exception:
        pass

    # Method 4 — pypdf (newer fork)
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(raw))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n".join(pages).strip()
        if text:
            return text
    except Exception:
        pass

    # Method 5 — raw byte scan (last resort, extracts readable strings)
    try:
        text = raw.decode("latin-1", errors="ignore")
        # Extract printable ASCII runs longer than 4 chars
        import re
        chunks = re.findall(r"[ -~]{4,}", text)
        readable = " ".join(chunks)
        if len(readable) > 100:
            return f"[PDF text extracted via fallback method]\n\n{readable[:15000]}"
    except Exception:
        pass

    return (
        "[⚠️ Could not extract PDF text. "
        "Please install pdfplumber: pip install pdfplumber]\n"
        "You can still ask general questions."
    )


def _read_docx(raw: bytes) -> str:
    """Try every available DOCX reader."""
    # Method 1 — python-docx
    try:
        import docx
        doc = docx.Document(io.BytesIO(raw))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception:
        pass

    # Method 2 — docx2txt
    try:
        import docx2txt
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name
        text = docx2txt.process(tmp_path)
        os.unlink(tmp_path)
        if text and text.strip():
            return text.strip()
    except Exception:
        pass

    # Method 3 — unzip and read XML (DOCX is a zip file)
    try:
        import zipfile, re
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            if "word/document.xml" in z.namelist():
                xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
                # Strip XML tags, extract text
                text = re.sub(r"<[^>]+>", " ", xml)
                text = re.sub(r"\s+", " ", text).strip()
                if text:
                    return f"[DOCX text extracted via XML fallback]\n\n{text[:15000]}"
    except Exception:
        pass

    return (
        "[⚠️ Could not extract DOCX text. "
        "Please install python-docx: pip install python-docx]\n"
        "You can still ask general questions."
    )


def _extract_text_from_file(uploaded) -> str:
    """Extract readable text from common file types — Lightning.ai safe."""
    fname = uploaded.name.lower()
    raw   = uploaded.read()

    # ── Plain text formats ────────────────────────────────────────────────
    if any(fname.endswith(ext) for ext in (".txt", ".md", ".csv", ".log", ".json", ".yaml", ".yml", ".html", ".htm", ".xml", ".py", ".js", ".ts", ".jsx", ".tsx", ".css", ".sh")):
        for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                return raw.decode(enc)
            except Exception:
                continue
        return raw.decode("utf-8", errors="replace")

    # ── PDF ───────────────────────────────────────────────────────────────
    if fname.endswith(".pdf"):
        return _read_pdf(raw)

    # ── DOCX ─────────────────────────────────────────────────────────────
    if fname.endswith(".docx") or fname.endswith(".doc"):
        return _read_docx(raw)

    # ── Excel / spreadsheet ───────────────────────────────────────────────
    if any(fname.endswith(ext) for ext in (".xlsx", ".xls", ".ods")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"=== Sheet: {sheet} ===")
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip():
                        lines.append(row_text)
            return "\n".join(lines)
        except Exception:
            pass
        try:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(raw))
            return df.to_string(index=False)
        except Exception:
            pass
        return "[⚠️ Could not read spreadsheet. Install openpyxl: pip install openpyxl]"

    # ── Images ───────────────────────────────────────────────────────────
    if any(fname.endswith(ext) for ext in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff")):
        b64 = base64.b64encode(raw).decode()
        ext = fname.split(".")[-1]
        return f"[IMAGE:{ext}:{b64}]"

    # ── Fallback: try UTF-8 ───────────────────────────────────────────────
    try:
        return raw.decode("utf-8")
    except Exception:
        return "[Binary file — cannot extract readable text]"


def _build_file_system_prompt(file_content: str, file_name: str) -> str:
    if file_content.startswith("[IMAGE:"):
        return (
            f"The user has uploaded an image file named '{file_name}'. "
            "Based on the filename, describe what this image likely contains and help the user with any related questions. "
            "You are NEXA Chatbot powered by LLaMA 3 — answer naturally and helpfully."
        )
    if file_content.startswith("[⚠️"):
        return (
            f"The user tried to upload '{file_name}' but text extraction failed. "
            "Inform them politely about the issue and offer help based on the filename alone. "
            "Suggest they install required libraries."
        )
    snippet = file_content[:12000]
    return (
        f"The user has uploaded a file named '{file_name}'. "
        f"Here is its content:\n\n---\n{snippet}\n---\n\n"
        "Answer any questions the user has about this file. "
        "Be thorough, accurate, and helpful. You are NEXA Chatbot powered by LLaMA 3."
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  GLOBAL CSS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CSS = """
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=JetBrains+Mono:wght@300;400;500;600&family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">

<style>
/* ── design tokens ── */
:root{
  --void:#030508; --surf:#07080f; --card:#0c0e1c; --inp:#111328; --hover:#161930; --lift:#1e2240;
  --bdr:rgba(255,255,255,.065); --bdr2:rgba(255,255,255,.12);
  --v1:#7c3aed; --v2:#a78bfa; --v3:#ddd6fe;
  --c1:#0891b2; --c2:#22d3ee; --c3:#a5f3fc;
  --acc:#f59e0b; --err:#ef4444; --ok:#10b981;
  --t1:#eef0ff; --t2:#8892b0; --t3:#343856;
  --display:'Syne',sans-serif; --body:'Inter',sans-serif; --mono:'JetBrains Mono',monospace;
  --r-xs:5px; --r-sm:10px; --r-md:14px; --r-lg:18px; --r-xl:26px; --r-2xl:36px;
  --shadow-lg:0 24px 80px rgba(0,0,0,.7);
}

/* ── shell reset ── */
html,body,[data-testid="stAppViewContainer"],[data-testid="stMain"],.main{
  background:var(--void)!important; color:var(--t1)!important; font-family:var(--body)!important;}
.main .block-container{
  background:transparent!important; padding-top:0!important;
  padding-bottom:7rem!important; max-width:1440px!important;}
#MainMenu,footer,header,[data-testid="stToolbar"],[data-testid="stDecoration"]{
  visibility:hidden!important; height:0!important;}
[data-testid="stAppViewBlockContainer"]{padding-top:.4rem!important;}
*{box-sizing:border-box;}

/* ── animated aurora ── */
#aurora{position:fixed;inset:0;z-index:0;pointer-events:none;overflow:hidden;background:var(--void);}
.ao{position:absolute;border-radius:50%;filter:blur(130px);will-change:transform;animation:drift linear infinite;}
.ao1{width:900px;height:900px;background:rgba(124,58,237,.18);top:-25%;left:-18%;animation-duration:28s;}
.ao2{width:700px;height:700px;background:rgba(8,145,178,.14);top:50%;right:-20%;animation-duration:35s;animation-delay:-12s;}
.ao3{width:600px;height:600px;background:rgba(167,139,250,.15);bottom:-15%;left:30%;animation-duration:22s;animation-delay:-6s;}
.ao4{width:500px;height:500px;background:rgba(245,158,11,.08);top:70%;left:8%; animation-duration:40s;animation-delay:-20s;}
.ao5{width:380px;height:380px;background:rgba(16,185,129,.09);top:5%; right:22%;animation-duration:31s;animation-delay:-15s;}
@keyframes drift{
  0%  {transform:translate(0,0)   scale(1);}
  25% {transform:translate(80px,-60px) scale(1.09);}
  50% {transform:translate(-50px,70px) scale(.92);}
  75% {transform:translate(55px,30px) scale(1.06);}
  100%{transform:translate(0,0)   scale(1);}
}
#aurora::after{
  content:'';position:absolute;inset:0;
  background-image:radial-gradient(circle,rgba(255,255,255,.028) 1px,transparent 1px);
  background-size:44px 44px;}

/* ── NEXA brand logo ── */
.nexa-logo{
  font-family:var(--display);font-weight:800;
  font-size:clamp(1.4rem,3vw,2.1rem);letter-spacing:-.04em;
  background:linear-gradient(135deg,#a78bfa 0%,#22d3ee 55%,#f59e0b 100%);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  background-clip:text;line-height:1;display:inline-block;}
.nexa-sub{
  font-family:var(--mono);font-size:.56rem;color:var(--t3);
  letter-spacing:.22em;text-transform:uppercase;margin-top:3px;}

/* ── inputs ── */
[data-testid="stTextInput"]>div,[data-testid="stTextInput"]{width:100%!important;}
[data-testid="stTextInput"] input{
  width:100%!important;background:var(--inp)!important;color:var(--t1)!important;
  border:1.5px solid var(--bdr)!important;border-radius:var(--r-md)!important;
  font-family:var(--mono)!important;font-size:.85rem!important;
  padding:.8rem 1rem!important;outline:none!important;
  transition:border-color .2s,box-shadow .2s!important;}
[data-testid="stTextInput"] input:focus{
  border-color:var(--v1)!important;box-shadow:0 0 0 3px rgba(124,58,237,.2)!important;}
[data-testid="stTextInput"] input::placeholder{color:var(--t3)!important;}

/* ── chat input ── */
[data-testid="stChatInput"]{
  background:rgba(6,7,15,.97)!important;border-top:1px solid var(--bdr)!important;
  backdrop-filter:blur(32px)!important;}
[data-testid="stChatInput"] textarea{
  background:var(--inp)!important;color:var(--t1)!important;
  font-family:var(--mono)!important;font-size:.85rem!important;
  border:1.5px solid var(--bdr)!important;border-radius:var(--r-md)!important;}
[data-testid="stChatInput"] textarea:focus{
  border-color:var(--v1)!important;box-shadow:0 0 0 3px rgba(124,58,237,.18)!important;}

/* ── file uploader ── */
[data-testid="stFileUploader"]{
  background:var(--inp)!important;border:1.5px dashed rgba(124,58,237,.35)!important;
  border-radius:var(--r-lg)!important;padding:.4rem!important;}
[data-testid="stFileUploader"]:hover{border-color:var(--v1)!important;}

/* ── main button ── */
.stButton>button{
  background:linear-gradient(135deg,var(--v1),#6d28d9)!important;
  color:#fff!important;border:none!important;border-radius:var(--r-md)!important;
  font-family:var(--display)!important;font-weight:700!important;font-size:.9rem!important;
  letter-spacing:.02em!important;padding:.8rem 2rem!important;width:100%!important;
  transition:all .22s ease!important;box-shadow:0 6px 28px rgba(124,58,237,.4)!important;}
.stButton>button:hover{
  transform:translateY(-2px)!important;filter:brightness(1.12)!important;
  box-shadow:0 14px 44px rgba(124,58,237,.6)!important;}
.stButton>button:active{transform:translateY(0)!important;}

/* ── sidebar ── */
[data-testid="stSidebar"]{
  background:linear-gradient(180deg,#08090f 0%,#060710 100%)!important;
  border-right:1px solid var(--bdr)!important;}
[data-testid="stSidebar"] *{color:var(--t1)!important;}
[data-testid="stSidebar"]>div{padding-top:0!important;}

[data-testid="stSidebar"] .stButton>button{
  background:transparent!important;border:1px solid transparent!important;
  color:var(--t2)!important;font-family:var(--body)!important;font-weight:500!important;
  font-size:.82rem!important;text-align:left!important;letter-spacing:0!important;
  box-shadow:none!important;border-radius:var(--r-sm)!important;
  padding:.52rem .9rem!important;margin:.04rem .5rem!important;
  width:calc(100% - 1rem)!important;transition:all .16s!important;}
[data-testid="stSidebar"] .stButton>button:hover{
  background:var(--hover)!important;border-color:var(--bdr2)!important;
  color:var(--t1)!important;transform:none!important;box-shadow:none!important;}
.sb-active .stButton>button{
  background:rgba(124,58,237,.14)!important;
  border-color:rgba(124,58,237,.32)!important;
  color:var(--v2)!important;}

.sb-danger .stButton>button{
  color:#f87171!important;border-color:rgba(239,68,68,.18)!important;
  background:rgba(239,68,68,.06)!important;}
.sb-danger .stButton>button:hover{
  background:rgba(239,68,68,.14)!important;border-color:rgba(239,68,68,.36)!important;}

/* ── sidebar components ── */
.sb-header{
  padding:1.1rem 1rem .9rem;border-bottom:1px solid var(--bdr);
  background:linear-gradient(160deg,rgba(124,58,237,.07),transparent);}
.sb-logo-row{display:flex;align-items:center;gap:.65rem;}
.sb-ico{
  width:34px;height:34px;border-radius:9px;
  background:linear-gradient(135deg,var(--v1),var(--c1));
  display:flex;align-items:center;justify-content:center;font-size:.95rem;flex-shrink:0;
  box-shadow:0 4px 18px rgba(124,58,237,.45);}

.sb-user{
  margin:.65rem .7rem;padding:.75rem .9rem;
  background:var(--card);border:1px solid var(--bdr);border-radius:var(--r-md);
  display:flex;align-items:center;gap:.7rem;}
.sb-ava{width:38px;height:38px;border-radius:9px;object-fit:cover;flex-shrink:0;
        border:2px solid rgba(124,58,237,.5);}
.sb-uname{font-weight:700;font-size:.82rem;line-height:1.2;font-family:var(--display);}
.sb-urole{font-family:var(--mono);font-size:.58rem;color:var(--t2)!important;margin-top:1px;}

.sb-lbl{
  font-family:var(--mono);font-size:.52rem;color:var(--t3)!important;
  text-transform:uppercase;letter-spacing:.22em;padding:.55rem 1rem .18rem;}
.sb-divider{border:none;border-top:1px solid var(--bdr);margin:.4rem .7rem;}

.sb-status{
  margin:0 .7rem .35rem;background:var(--card);
  border:1px solid var(--bdr);border-radius:var(--r-sm);padding:.55rem .8rem;}
.sb-srow{
  display:flex;justify-content:space-between;align-items:center;
  font-family:var(--mono);font-size:.68rem;color:var(--t2)!important;padding:.12rem 0;}

.pill{display:inline-flex;align-items:center;gap:.25rem;
      font-family:var(--mono);font-size:.56rem;font-weight:600;
      padding:.12rem .48rem;border-radius:999px;letter-spacing:.06em;}
.pill-ok {background:rgba(16,185,129,.12);color:#34d399!important;border:1px solid rgba(16,185,129,.28);}
.pill-err{background:rgba(239,68,68,.1); color:#f87171!important;border:1px solid rgba(239,68,68,.24);}
.pill-dot{width:5px;height:5px;border-radius:50%;flex-shrink:0;}
.dot-ok {background:#34d399;animation:pulse 2s infinite;}
.dot-err{background:#f87171;}
@keyframes pulse{0%,100%{opacity:1;transform:scale(1)}50%{opacity:.5;transform:scale(.7)}}

.sb-stats{display:grid;grid-template-columns:1fr 1fr;gap:.4rem;margin:0 .7rem .35rem;}
.sb-stat{background:var(--card);border:1px solid var(--bdr);
         border-radius:var(--r-sm);padding:.55rem .65rem;text-align:center;}
.sb-sval{font-family:var(--mono);font-size:1.15rem;font-weight:700;color:var(--v2)!important;line-height:1;}
.sb-slbl{font-family:var(--mono);font-size:.48rem;color:var(--t3)!important;
         text-transform:uppercase;letter-spacing:.1em;margin-top:2px;}

/* ── alerts ── */
[data-testid="stAlert"]{
  background:rgba(239,68,68,.07)!important;border:1px solid rgba(239,68,68,.22)!important;
  border-radius:var(--r-sm)!important;font-family:var(--mono)!important;font-size:.8rem!important;}
[data-testid="stWarning"]{
  background:rgba(245,158,11,.06)!important;border:1px solid rgba(245,158,11,.2)!important;
  border-radius:var(--r-sm)!important;}
[data-testid="stInfo"]{
  background:rgba(124,58,237,.07)!important;border:1px solid rgba(124,58,237,.22)!important;
  border-radius:var(--r-sm)!important;}
[data-testid="stSuccess"]{
  background:rgba(16,185,129,.06)!important;border:1px solid rgba(16,185,129,.22)!important;
  border-radius:var(--r-sm)!important;}
pre,code{
  font-family:var(--mono)!important;background:var(--inp)!important;
  border:1px solid var(--bdr)!important;border-radius:var(--r-xs)!important;
  color:var(--c2)!important;}
[data-testid="stVerticalBlock"]{gap:.18rem!important;}
[data-testid="stMarkdownContainer"] p{margin:0;}
::-webkit-scrollbar{width:3px;height:3px;}
::-webkit-scrollbar-track{background:transparent;}
::-webkit-scrollbar-thumb{background:var(--lift);border-radius:3px;}
hr{border:none;border-top:1px solid var(--bdr);margin:.6rem 0;}

/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   LOGIN PAGE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
.login-page{
  position:relative;z-index:1;
  min-height:95vh;display:flex;flex-direction:column;
  align-items:center;justify-content:center;padding:2rem 1rem;}

.login-brand-banner{text-align:center;margin-bottom:1.6rem;}
.login-brand-title{
  font-family:var(--display);font-weight:800;
  font-size:clamp(2rem,5vw,3.2rem);letter-spacing:-.05em;
  background:linear-gradient(135deg,#a78bfa 0%,#22d3ee 55%,#f59e0b 100%);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  background-clip:text;line-height:1.05;margin-bottom:.35rem;}
.login-brand-sub{
  font-family:var(--mono);font-size:.64rem;color:var(--t3);
  letter-spacing:.22em;text-transform:uppercase;}

.uc-label{
  font-family:var(--mono);font-size:.54rem;color:var(--t3);
  text-transform:uppercase;letter-spacing:.2em;
  margin-bottom:.5rem;display:flex;align-items:center;gap:.35rem;}
.uc-grid{display:grid;grid-template-columns:1fr 1fr;gap:.6rem;margin-bottom:1rem;width:100%;}
.uc{
  background:rgba(12,14,28,.88);border:1.5px solid var(--bdr);
  border-radius:var(--r-md);padding:.85rem .8rem;text-align:center;
  transition:all .22s ease;position:relative;overflow:hidden;backdrop-filter:blur(20px);}
.uc::after{content:'';position:absolute;bottom:0;left:0;right:0;height:2px;opacity:0;transition:opacity .22s;}
.uc-admin::after{background:linear-gradient(90deg,var(--v1),var(--v2));}
.uc-demo::after {background:linear-gradient(90deg,var(--c1),var(--c2));}
.uc:hover{transform:translateY(-3px);border-color:rgba(255,255,255,.14);}
.uc:hover::after{opacity:1;}
.uc-admin:hover{box-shadow:0 14px 38px rgba(124,58,237,.2);border-color:rgba(124,58,237,.35);}
.uc-demo:hover {box-shadow:0 14px 38px rgba(8,145,178,.18);border-color:rgba(8,145,178,.32);}
.uc-avatar{width:48px;height:48px;border-radius:12px;object-fit:cover;display:block;margin:0 auto .6rem;}
.uc-admin .uc-avatar{border:2px solid rgba(124,58,237,.55);}
.uc-demo  .uc-avatar{border:2px solid rgba(8,145,178,.45);}
.uc-name{font-family:var(--mono);font-size:.78rem;font-weight:600;margin-bottom:.18rem;}
.uc-pass{font-family:var(--mono);font-size:.64rem;color:var(--t3);margin-bottom:.4rem;}
.uc-badge{
  display:inline-block;font-family:var(--mono);font-size:.53rem;font-weight:700;
  padding:.11rem .5rem;border-radius:999px;letter-spacing:.08em;}
.badge-admin{background:rgba(124,58,237,.15);color:var(--v2)!important;border:1px solid rgba(124,58,237,.35);}
.badge-demo {background:rgba(8,145,178,.12);color:var(--c2)!important; border:1px solid rgba(8,145,178,.3);}

.login-card{
  width:100%;background:rgba(10,11,22,.84);
  border:1px solid rgba(255,255,255,.08);border-radius:var(--r-2xl);
  padding:2rem 2.2rem 1.8rem;
  box-shadow:0 40px 120px rgba(0,0,0,.8),0 0 0 1px rgba(255,255,255,.03) inset,0 0 90px rgba(124,58,237,.06);
  backdrop-filter:blur(48px) saturate(1.5);position:relative;overflow:hidden;}
.login-card::before{
  content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,transparent,var(--v1),var(--c1),var(--acc),transparent);
  background-size:300% 100%;animation:shimmer 5s linear infinite;}
@keyframes shimmer{0%{background-position:0%}100%{background-position:300%}}

.brand-ico{
  width:58px;height:58px;border-radius:16px;
  background:linear-gradient(135deg,var(--v1),var(--c1));
  display:flex;align-items:center;justify-content:center;
  font-size:1.6rem;margin:0 auto .75rem;
  box-shadow:0 10px 34px rgba(124,58,237,.5),0 0 0 10px rgba(124,58,237,.07);
  animation:brand-pulse 3.5s ease-in-out infinite;}
@keyframes brand-pulse{
  0%,100%{box-shadow:0 10px 34px rgba(124,58,237,.5),0 0 0 10px rgba(124,58,237,.07);}
  50%    {box-shadow:0 10px 48px rgba(124,58,237,.72),0 0 0 18px rgba(124,58,237,.04);}}

.sec-strip{
  display:flex;align-items:center;gap:.6rem;
  background:rgba(16,185,129,.07);border:1px solid rgba(16,185,129,.18);
  border-radius:var(--r-sm);padding:.55rem .85rem;margin-bottom:1.3rem;
  font-family:var(--mono);font-size:.65rem;color:#6ee7b7;}
.sec-dot{
  width:7px;height:7px;border-radius:50%;
  background:var(--ok);box-shadow:0 0 7px var(--ok);animation:pulse 2s infinite;flex-shrink:0;}

.f-label{
  font-family:var(--mono);font-size:.59rem;font-weight:500;
  color:var(--t2)!important;text-transform:uppercase;letter-spacing:.14em;
  margin-bottom:.32rem;margin-top:.85rem;display:flex;align-items:center;gap:.3rem;}

/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   POST-LOGIN TOPBAR
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
.topbar{
  display:flex;align-items:center;justify-content:space-between;
  padding:.75rem 0 1rem;border-bottom:1px solid var(--bdr);margin-bottom:1.25rem;
  position:relative;z-index:1;}
.topbar-left{display:flex;align-items:center;gap:.75rem;}
.tb-ico{width:40px;height:40px;border-radius:11px;display:flex;align-items:center;justify-content:center;font-size:1.05rem;}
.t-chat {background:rgba(124,58,237,.13);border:1px solid rgba(124,58,237,.24);}
.t-admin{background:rgba(245,158,11,.11); border:1px solid rgba(245,158,11,.22);}
.t-model{background:rgba(8,145,178,.11);  border:1px solid rgba(8,145,178,.22);}
.tb-title{font-size:1.1rem;font-weight:800;letter-spacing:-.025em;margin:0;font-family:var(--display);}
.tb-sub  {font-family:var(--mono);font-size:.6rem;color:var(--t3);margin-top:1px;}

.topbar-right{display:flex;align-items:center;gap:.7rem;}
.live-badge{
  font-family:var(--mono);font-size:.57rem;font-weight:700;
  padding:.18rem .6rem;border-radius:999px;letter-spacing:.08em;
  background:rgba(16,185,129,.1);color:#34d399!important;
  border:1px solid rgba(16,185,129,.25);display:flex;align-items:center;gap:.28rem;}
.live-dot{width:5px;height:5px;border-radius:50%;background:#34d399;animation:pulse 2s infinite;}
.page-nexa{
  font-family:var(--display);font-weight:800;font-size:1.55rem;letter-spacing:-.04em;
  background:linear-gradient(135deg,#a78bfa,#22d3ee);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}

/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   OFFLINE BANNER
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
.offline-banner{
  display:flex;align-items:flex-start;gap:1rem;
  background:rgba(239,68,68,.07);border:1px solid rgba(239,68,68,.28);
  border-left:4px solid var(--err);border-radius:var(--r-md);
  padding:1rem 1.2rem;margin-bottom:1.25rem;animation:fade-slide .4s ease;}
@keyframes fade-slide{from{opacity:0;transform:translateY(-8px)}to{opacity:1;transform:none}}
.ob-icon{font-size:1.4rem;flex-shrink:0;margin-top:.05rem;}
.ob-title{font-weight:800;font-size:.94rem;color:#fca5a5;margin-bottom:.2rem;font-family:var(--display);}
.ob-body {font-family:var(--mono);font-size:.73rem;color:#f87171;line-height:1.7;}
.ob-cmd{
  display:inline-block;background:rgba(239,68,68,.11);
  border:1px solid rgba(239,68,68,.28);border-radius:var(--r-xs);
  font-family:var(--mono);font-size:.73rem;color:#fca5a5;padding:.1rem .55rem;margin:.1rem .1rem 0 0;}
.ob-steps{margin-top:.45rem;display:flex;flex-wrap:wrap;gap:.3rem;}

/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   FILE UPLOAD PANEL
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
.file-panel{
  background:rgba(124,58,237,.06);border:1px solid rgba(124,58,237,.2);
  border-radius:var(--r-md);padding:.9rem 1.1rem;margin-bottom:1rem;
  display:flex;align-items:center;gap:.85rem;animation:fade-slide .3s ease;}
.file-ico{font-size:1.5rem;flex-shrink:0;}
.file-name{font-family:var(--mono);font-size:.8rem;font-weight:600;color:var(--v2);}
.file-meta{font-family:var(--mono);font-size:.62rem;color:var(--t3);margin-top:2px;}

/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   CHAT MESSAGES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
.chat-empty{
  text-align:center;padding:3rem 2rem;
  display:flex;flex-direction:column;align-items:center;position:relative;z-index:1;}
.chat-orb{
  width:84px;height:84px;border-radius:24px;margin-bottom:1.2rem;font-size:2.1rem;
  background:linear-gradient(135deg,rgba(124,58,237,.14),rgba(8,145,178,.08));
  border:1px solid rgba(124,58,237,.22);
  display:flex;align-items:center;justify-content:center;
  box-shadow:0 0 70px rgba(124,58,237,.14);animation:float 4s ease-in-out infinite;}
@keyframes float{0%,100%{transform:translateY(0)rotate(0)}50%{transform:translateY(-9px)rotate(1.5deg)}}
.chat-title{font-size:1.15rem;font-weight:800;letter-spacing:-.02em;margin-bottom:.3rem;font-family:var(--display);}
.chat-sub  {font-family:var(--mono);font-size:.65rem;color:var(--t3);margin-bottom:1.5rem;}
.chip-grid {display:flex;flex-wrap:wrap;gap:.38rem;justify-content:center;max-width:580px;}
.chip{
  background:var(--card);border:1px solid var(--bdr);border-radius:var(--r-sm);
  padding:.38rem .82rem;font-family:var(--mono);font-size:.68rem;color:var(--t2)!important;
  display:inline-flex;align-items:center;gap:.32rem;transition:all .15s;cursor:default;}
.chip:hover{border-color:var(--v1);color:var(--t1)!important;background:var(--hover);}

.msg-row{margin:.42rem 0;animation:msg-pop .28s cubic-bezier(.34,1.56,.64,1);position:relative;z-index:1;}
@keyframes msg-pop{from{opacity:0;transform:translateY(10px)scale(.97)}to{opacity:1;transform:none}}

.msg-user{
  background:linear-gradient(145deg,#0e1240,#141a58);
  border:1px solid rgba(124,58,237,.26);border-radius:4px 18px 18px 18px;
  padding:.85rem 1.05rem;font-family:var(--mono);font-size:.82rem;line-height:1.8;
  max-width:74%;margin-left:auto;box-shadow:0 4px 22px rgba(0,0,0,.4);}
.msg-ai{
  background:linear-gradient(145deg,#070f15,#091621);
  border:1px solid rgba(8,145,178,.17);border-radius:18px 18px 18px 4px;
  padding:.85rem 1.05rem;font-family:var(--mono);font-size:.82rem;line-height:1.8;
  max-width:74%;box-shadow:0 4px 22px rgba(0,0,0,.4);}

.msg-meta{
  display:flex;align-items:center;gap:.32rem;margin-bottom:.38rem;
  font-family:var(--mono);font-size:.53rem;font-weight:700;
  letter-spacing:.16em;text-transform:uppercase;}
.meta-u{color:#a78bfa!important;} .meta-a{color:#22d3ee!important;}
.mdot  {width:4px;height:4px;border-radius:50%;}
.mdu   {background:#a78bfa;box-shadow:0 0 5px #a78bfa;}
.mda   {background:#22d3ee;box-shadow:0 0 5px #22d3ee;}
.cursor{color:var(--v1);animation:blink .75s step-end infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:0}}

/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   GLASS CARDS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
.gc{
  background:linear-gradient(160deg,var(--card),var(--hover));
  border:1px solid var(--bdr);border-radius:var(--r-lg);
  padding:1.3rem 1.5rem;box-shadow:0 8px 40px rgba(0,0,0,.55);
  position:relative;overflow:hidden;}
.gc::before{content:'';position:absolute;top:0;left:0;right:0;height:1px;
  background:linear-gradient(90deg,transparent,rgba(255,255,255,.05),transparent);}

.pphoto{
  width:76px;height:76px;border-radius:18px;object-fit:cover;
  display:block;margin:0 auto .75rem;border:3px solid var(--v1);
  box-shadow:0 0 0 6px rgba(124,58,237,.14),0 12px 34px rgba(0,0,0,.5);}

.ptab{width:100%;border-collapse:collapse;font-family:var(--mono);font-size:.77rem;}
.ptab td{padding:.55rem .35rem;vertical-align:middle;}
.ptab tr{border-bottom:1px solid var(--bdr);}
.ptab tr:last-child{border-bottom:none;}
.ptab .k{color:var(--t3)!important;font-size:.67rem;text-transform:uppercase;letter-spacing:.08em;width:38%;}
.ptab .v{color:var(--t1)!important;font-weight:500;}

.mini-grid{display:grid;grid-template-columns:1fr 1fr;gap:.45rem;margin-top:.85rem;}
.mini-cell{background:var(--hover);border:1px solid var(--bdr);border-radius:var(--r-sm);padding:.62rem;text-align:center;}
.mini-val {font-family:var(--mono);font-size:1.2rem;font-weight:800;line-height:1;}
.mini-lbl {font-family:var(--mono);font-size:.49rem;color:var(--t3)!important;text-transform:uppercase;letter-spacing:.1em;margin-top:2px;}

.metric{
  background:var(--card);border:1px solid var(--bdr);border-radius:var(--r-md);
  padding:1rem 1.1rem;position:relative;overflow:hidden;transition:all .2s;}
.metric:hover{border-color:rgba(255,255,255,.11);transform:translateY(-2px);}
.metric::after{content:'';position:absolute;bottom:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,var(--v1),var(--c1));opacity:0;transition:opacity .2s;}
.metric:hover::after{opacity:1;}
.metric-ico{position:absolute;top:.8rem;right:.85rem;font-size:.9rem;opacity:.18;}
.metric-lbl{font-family:var(--mono);font-size:.54rem;color:var(--t3)!important;text-transform:uppercase;letter-spacing:.15em;margin-bottom:.35rem;}
.metric-val{font-family:var(--display);font-size:.92rem;font-weight:700;}

.ok-badge{
  display:inline-flex;align-items:center;gap:.26rem;padding:.12rem .55rem;border-radius:999px;
  font-family:var(--mono);font-size:.57rem;font-weight:700;letter-spacing:.06em;
  background:rgba(16,185,129,.1);color:#34d399!important;border:1px solid rgba(16,185,129,.25);}
.ok-dot{width:5px;height:5px;border-radius:50%;background:#34d399;animation:pulse 2s infinite;flex-shrink:0;}

.spec-row{
  display:flex;justify-content:space-between;align-items:center;
  padding:.55rem 0;border-bottom:1px solid var(--bdr);
  font-family:var(--mono);font-size:.76rem;}
.spec-row:last-child{border-bottom:none;}
.spec-k{color:var(--t3);font-size:.66rem;text-transform:uppercase;letter-spacing:.1em;}
.spec-v{color:var(--c2);font-weight:600;}

.act-item{display:flex;align-items:flex-start;gap:.7rem;padding:.55rem 0;border-bottom:1px solid var(--bdr);}
.act-item:last-child{border-bottom:none;}
.act-dot{width:8px;height:8px;border-radius:50%;background:var(--v2);margin-top:.28rem;flex-shrink:0;}
.act-text{font-family:var(--mono);font-size:.74rem;color:var(--t2);}
.act-time{font-family:var(--mono);font-size:.58rem;color:var(--t3);margin-top:2px;}
</style>

<!-- Aurora background -->
<div id="aurora">
  <div class="ao ao1"></div>
  <div class="ao ao2"></div>
  <div class="ao ao3"></div>
  <div class="ao ao4"></div>
  <div class="ao ao5"></div>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
"""

def _inject_styles():
    st.markdown(CSS, unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  REUSABLE COMPONENTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _offline_banner():
    st.markdown("""
<div class="offline-banner">
  <div class="ob-icon">⚡</div>
  <div>
    <div class="ob-title">Ollama Engine Offline</div>
    <div class="ob-body">
      The local LLaMA 3 engine isn't running. Chat is disabled until Ollama is started.<br>
      Run these commands in your terminal:
      <div class="ob-steps">
        <span class="ob-cmd">bash setup.sh</span>
        <span class="ob-cmd">bash run.sh</span>
      </div>
    </div>
  </div>
</div>""", unsafe_allow_html=True)


def _topbar(icon, ico_cls, title, subtitle, badge_html=""):
    full = badge_html or '<div class="live-badge"><span class="live-dot"></span>LOCAL</div>'
    st.markdown(f"""
<div class="topbar">
  <div class="topbar-left">
    <div class="tb-ico {ico_cls}">{icon}</div>
    <div>
      <div class="tb-title">{title}</div>
      <div class="tb-sub">{subtitle}</div>
    </div>
  </div>
  <div class="topbar-right">
    <div class="page-nexa">NEXA CHATBOT</div>
    {full}
  </div>
</div>""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  LOGIN PAGE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def render_login():
    adm = AVATAR.get("admin", "")
    dem = AVATAR.get("demo",  "")

    st.markdown('<div class="login-page">', unsafe_allow_html=True)
    _, mid, _ = st.columns([1, 1.3, 1])

    with mid:
        st.markdown("""
        <div class="login-brand-banner">
          <div class="login-brand-title">NEXA CHATBOT</div>
          <div class="login-brand-sub">LLaMA 3 · 8B · Fully Local · Zero Cloud</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class="uc-label">
          <i class="bi bi-people-fill"></i>&nbsp; Demo Accounts
        </div>
        <div class="uc-grid">
          <div class="uc uc-admin">
            <img src="{adm}" class="uc-avatar" alt="Admin">
            <div class="uc-name">admin</div>
            <div class="uc-pass">admin123</div>
            <span class="uc-badge badge-admin">Super Admin</span>
          </div>
          <div class="uc uc-demo">
            <img src="{dem}" class="uc-avatar" alt="Demo">
            <div class="uc-name">demo</div>
            <div class="uc-pass">demo2024</div>
            <span class="uc-badge badge-demo">Analyst</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class="login-card">
          <div style="text-align:center;margin-bottom:1.2rem">
            <div class="brand-ico">⚡</div>
            <div style="font-family:var(--display);font-weight:800;font-size:1.35rem;
                        letter-spacing:-.03em;margin-bottom:.12rem">Sign In</div>
            <div style="font-family:var(--mono);font-size:.6rem;color:var(--t3);
                        letter-spacing:.18em;text-transform:uppercase">
              Access your local AI workspace
            </div>
          </div>
          <div class="sec-strip">
            <span class="sec-dot"></span>
            End-to-end local &nbsp;·&nbsp; Zero telemetry &nbsp;·&nbsp; No data leaves this machine
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="f-label"><i class="bi bi-person-fill"></i>&nbsp; Username</div>',
                    unsafe_allow_html=True)
        username = st.text_input("u", placeholder="Enter username",
                                 key="li_user", label_visibility="collapsed")

        st.markdown('<div class="f-label"><i class="bi bi-shield-lock-fill"></i>&nbsp; Password</div>',
                    unsafe_allow_html=True)
        password = st.text_input("p", placeholder="Enter password",
                                 type="password", key="li_pass",
                                 label_visibility="collapsed")

        if st.session_state.login_error:
            st.error(st.session_state.login_error)

        st.markdown("<div style='height:.6rem'></div>", unsafe_allow_html=True)

        if st.button("Sign In →", use_container_width=True, key="btn_login"):
            if not username or not password:
                st.session_state.login_error = "Please fill in both fields."
                st.rerun()
            else:
                info = authenticate(username, password)
                if info:
                    st.session_state.logged_in    = True
                    st.session_state.username     = username
                    st.session_state.admin_info   = info
                    st.session_state.conversation = ConversationManager()
                    st.session_state.messages     = []
                    st.session_state.login_error  = ""
                    st.rerun()
                else:
                    st.session_state.login_error = "❌ Invalid credentials — check username and password."
                    st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SIDEBAR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def render_sidebar():
    info     = st.session_state.admin_info or {}
    conv     = st.session_state.conversation
    username = st.session_state.username or ""
    photo    = AVATAR.get(username, AVATAR.get("demo"))
    page     = st.session_state.page

    ollama_ok = check_ollama_running()
    model_ok  = check_model_available() if ollama_ok else False

    ob = (f'<span class="pill pill-ok"><span class="pill-dot dot-ok"></span>ONLINE</span>'
          if ollama_ok else
          f'<span class="pill pill-err"><span class="pill-dot dot-err"></span>OFFLINE</span>')
    mb = (f'<span class="pill pill-ok"><span class="pill-dot dot-ok"></span>LOADED</span>'
          if model_ok else
          f'<span class="pill pill-err"><span class="pill-dot dot-err"></span>MISSING</span>')

    stats = conv.stats if conv else {"messages": 0, "tokens_est": 0}
    fname = st.session_state.file_name

    with st.sidebar:
        st.markdown(f"""
        <div class="sb-header">
          <div class="sb-logo-row">
            <div class="sb-ico">⚡</div>
            <div>
              <div style="font-family:var(--display);font-weight:800;font-size:.92rem;
                          letter-spacing:-.02em;background:linear-gradient(135deg,#a78bfa,#22d3ee);
                          -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                          background-clip:text;">NEXA CHATBOT</div>
              <div style="font-family:var(--mono);font-size:.5rem;color:var(--t3);
                          letter-spacing:.2em;text-transform:uppercase;margin-top:1px;">
                LLaMA 3 · Local AI
              </div>
            </div>
          </div>
        </div>

        <div class="sb-user">
          <img src="{photo}" class="sb-ava" alt="{username}">
          <div>
            <div class="sb-uname">{info.get('full_name','—')}</div>
            <div class="sb-urole">{info.get('role','—')}</div>
          </div>
        </div>

        <div class="sb-lbl">System Status</div>
        <div class="sb-status">
          <div class="sb-srow"><span>Ollama Engine</span>{ob}</div>
          <div class="sb-srow" style="margin-top:.25rem">
            <span style="font-size:.66rem;overflow:hidden;text-overflow:ellipsis;
                         white-space:nowrap;max-width:54%">{MODEL_NAME}</span>{mb}
          </div>
        </div>
        """, unsafe_allow_html=True)

        if not ollama_ok:
            st.warning("Ollama offline — run `bash run.sh`", icon="⚠️")

        if fname:
            ext = fname.split(".")[-1].upper() if "." in fname else "FILE"
            st.markdown(f"""
            <div class="sb-lbl">Uploaded File</div>
            <div class="sb-status">
              <div class="sb-srow">
                <span>📎 {fname[:20]}{'…' if len(fname)>20 else ''}</span>
                <span class="pill pill-ok"><span class="pill-dot dot-ok"></span>{ext}</span>
              </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class="sb-lbl">Session Stats</div>
        <div class="sb-stats">
          <div class="sb-stat">
            <div class="sb-sval">{stats['messages']}</div>
            <div class="sb-slbl">Messages</div>
          </div>
          <div class="sb-stat">
            <div class="sb-sval" style="color:var(--c2)!important">{stats['tokens_est']}</div>
            <div class="sb-slbl">Tokens~</div>
          </div>
        </div>
        <div class="sb-lbl">Navigate</div>
        """, unsafe_allow_html=True)

        for nav_id, label in [("chat", "💬  Chat"), ("admin", "🧑‍💼  Admin Profile"), ("model", "🤖  Model Info")]:
            css_cls = "sb-active" if page == nav_id else ""
            st.markdown(f'<div class="{css_cls}">', unsafe_allow_html=True)
            if st.button(label, use_container_width=True, key=f"nav_{nav_id}"):
                st.session_state.page = nav_id; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        if st.button("🗑️  Clear Chat", use_container_width=True, key="nav_clear"):
            if conv: conv.clear()
            st.session_state.messages     = []
            st.session_state.file_content = None
            st.session_state.file_name    = None
            st.session_state.file_type    = None
            st.rerun()

        st.markdown('<hr class="sb-divider"><div class="sb-danger">', unsafe_allow_html=True)
        if st.button("🚪  Sign Out", use_container_width=True, key="nav_logout"):
            for k, v in _DEFAULTS.items():
                st.session_state[k] = v
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PAGE — CHAT (with file upload)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def page_chat():
    ollama_ok = check_ollama_running()

    _topbar("💬", "t-chat", "Chat", "Streaming · Multi-turn · File Q&A · Zero cloud")

    if not ollama_ok:
        _offline_banner()

    if st.session_state.conversation is None:
        st.session_state.conversation = ConversationManager()
    conv = st.session_state.conversation
    msgs = st.session_state.messages

    # ── File upload panel ────────────────────────────────────────────────
    with st.expander("📎  Upload a file to chat with it", expanded=bool(st.session_state.file_name)):
        st.markdown("""
        <div style="font-family:var(--mono);font-size:.67rem;color:var(--t3);margin-bottom:.5rem">
          Supported: <b style="color:var(--t2)">PDF · TXT · MD · DOCX · CSV · XLSX · PNG · JPG</b>
          — ask any question about its contents.
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Drop file here",
            type=["pdf", "txt", "md", "docx", "doc", "csv", "xlsx", "xls",
                  "json", "yaml", "yml", "py", "js", "html",
                  "png", "jpg", "jpeg", "webp", "gif"],
            label_visibility="collapsed",
            key="file_uploader",
        )

        if uploaded is not None:
            if uploaded.name != st.session_state.file_name:
                with st.spinner("🔍 Extracting file content…"):
                    content = _extract_text_from_file(uploaded)
                st.session_state.file_content = content
                st.session_state.file_name    = uploaded.name
                st.session_state.file_type    = uploaded.type
                if content.startswith("[⚠️"):
                    st.warning(f"⚠️ Partial extraction for **{uploaded.name}** — some content may be missing.")
                else:
                    st.success(f"✅ File loaded: **{uploaded.name}** — now ask anything about it!")

        if st.session_state.file_name:
            fname = st.session_state.file_name
            ext   = fname.split(".")[-1].upper() if "." in fname else "FILE"
            ico   = {
                "PDF":"📄","TXT":"📝","MD":"📝","DOCX":"📄","DOC":"📄",
                "CSV":"📊","XLSX":"📊","XLS":"📊","JSON":"📋","PY":"🐍",
                "JS":"📜","HTML":"🌐","PNG":"🖼️","JPG":"🖼️","JPEG":"🖼️",
                "WEBP":"🖼️","GIF":"🖼️",
            }.get(ext, "📎")
            st.markdown(f"""
            <div class="file-panel">
              <div class="file-ico">{ico}</div>
              <div>
                <div class="file-name">{fname}</div>
                <div class="file-meta">{ext} file · ready for Q&amp;A</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button("✕ Remove file", key="remove_file"):
                st.session_state.file_content = None
                st.session_state.file_name    = None
                st.session_state.file_type    = None
                st.rerun()

    # ── Empty state ──────────────────────────────────────────────────────
    if not msgs:
        st.markdown("""
        <div class="chat-empty">
          <div class="chat-orb">⚡</div>
          <div class="chat-title">NEXA is Ready</div>
          <div class="chat-sub">100% local · No internet · No logs · Upload files &amp; ask anything</div>
          <div class="chip-grid">
            <span class="chip">⚛️ Explain quantum computing</span>
            <span class="chip">🐍 Write a Python scraper</span>
            <span class="chip">🐞 Debug my code</span>
            <span class="chip">📄 Summarise uploaded file</span>
            <span class="chip">✉️ Draft an email</span>
            <span class="chip">🌍 Translate text</span>
            <span class="chip">📊 Analyse my CSV data</span>
            <span class="chip">🧮 Explain this math</span>
          </div>
        </div>""", unsafe_allow_html=True)
    else:
        for m in msgs:
            if m["role"] == "user":
                st.markdown(
                    f'<div class="msg-row" style="display:flex;justify-content:flex-end;padding-left:12%">'
                    f'<div class="msg-user">'
                    f'<div class="msg-meta meta-u"><span class="mdot mdu"></span>You</div>'
                    f'{m["content"]}</div></div>',
                    unsafe_allow_html=True)
            else:
                st.markdown(
                    f'<div class="msg-row" style="padding-right:12%">'
                    f'<div class="msg-ai">'
                    f'<div class="msg-meta meta-a"><span class="mdot mda"></span>NEXA · LLaMA 3</div>'
                    f'{m["content"]}</div></div>',
                    unsafe_allow_html=True)

    placeholder_text = "Ask about your file…" if st.session_state.file_name else "Message NEXA…"
    if not ollama_ok:
        placeholder_text = "⚠️ Ollama offline — start Ollama first"

    user_input = st.chat_input(
        placeholder_text,
        disabled=not ollama_ok,
        key="chat_input_box",
    )

    if user_input and user_input.strip() and ollama_ok:
        clean = user_input.strip()

        if st.session_state.file_content:
            file_prefix = (
                f"[Context from file '{st.session_state.file_name}']\n"
                f"{st.session_state.file_content[:10000]}\n\n"
                f"[User question]: "
            )
            augmented_input = file_prefix + clean
        else:
            augmented_input = clean

        st.markdown(
            f'<div class="msg-row" style="display:flex;justify-content:flex-end;padding-left:12%">'
            f'<div class="msg-user">'
            f'<div class="msg-meta meta-u"><span class="mdot mdu"></span>You</div>'
            f'{clean}</div></div>',
            unsafe_allow_html=True)

        placeholder = st.empty()
        tokens = []
        for tok in stream_response(conv, augmented_input):
            tokens.append(tok)
            placeholder.markdown(
                f'<div class="msg-row" style="padding-right:12%">'
                f'<div class="msg-ai">'
                f'<div class="msg-meta meta-a"><span class="mdot mda"></span>NEXA · LLaMA 3</div>'
                f'{"".join(tokens)}<span class="cursor">▌</span></div></div>',
                unsafe_allow_html=True)

        placeholder.empty()
        final = "".join(tokens)
        st.session_state.messages.append({"role": "user",      "content": clean})
        st.session_state.messages.append({"role": "assistant", "content": final})
        st.rerun()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PAGE — ADMIN PROFILE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def page_admin():
    info   = st.session_state.admin_info or {}
    user   = st.session_state.username   or "—"
    conv   = st.session_state.conversation
    sess   = conv.stats["session_start"] if conv else "—"
    msgs   = len(st.session_state.messages)
    tokens = conv.stats["tokens_est"] if conv else 0
    photo  = AVATAR.get(user, AVATAR.get("demo"))
    fname  = st.session_state.file_name

    _topbar("🧑‍💼", "t-admin", "Admin Profile", "Account · Session · Permissions",
            '<span class="ok-badge"><span class="ok-dot"></span>ACTIVE</span>')

    c1, c2 = st.columns([1, 2.2], gap="large")

    with c1:
        st.markdown(f"""
        <div class="gc" style="text-align:center">
          <img src="{photo}" class="pphoto" alt="{user}">
          <div style="font-family:var(--display);font-weight:800;font-size:1rem;letter-spacing:-.015em">
            {info.get('full_name','—')}</div>
          <div style="font-family:var(--mono);font-size:.63rem;color:var(--t2);margin-top:.15rem">
            {info.get('role','—')}</div>
          <hr>
          <div class="mini-grid">
            <div class="mini-cell">
              <div class="mini-val" style="color:var(--v2)">{msgs}</div>
              <div class="mini-lbl">Messages</div>
            </div>
            <div class="mini-cell">
              <div class="mini-val" style="color:var(--c2)">{tokens}</div>
              <div class="mini-lbl">Tokens~</div>
            </div>
          </div>
          <div class="mini-grid">
            <div class="mini-cell">
              <div class="mini-val" style="color:var(--acc);font-size:.9rem">
                {"1" if fname else "0"}
              </div>
              <div class="mini-lbl">File Loaded</div>
            </div>
            <div class="mini-cell">
              <div class="mini-val" style="color:#34d399;font-size:.9rem">ON</div>
              <div class="mini-lbl">AI Active</div>
            </div>
          </div>
          <div style="margin-top:.85rem">
            <span class="ok-badge"><span class="ok-dot"></span>ONLINE SESSION</span>
          </div>
        </div>""", unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
        <div class="gc" style="margin-bottom:.85rem">
          <div style="font-family:var(--mono);font-size:.54rem;color:var(--t3);
               text-transform:uppercase;letter-spacing:.17em;margin-bottom:.85rem">
            <i class="bi bi-person-badge"></i>&nbsp; Account Details
          </div>
          <table class="ptab">
            <tr>
              <td class="k">Username</td>
              <td class="v">
                <code style="background:var(--hover);padding:.1rem .48rem;
                     border-radius:5px;font-size:.73rem">{user}</code>
              </td>
            </tr>
            <tr><td class="k">Full Name</td>   <td class="v">{info.get('full_name','—')}</td></tr>
            <tr>
              <td class="k">Role</td>
              <td class="v">
                <span class="ok-badge" style="font-size:.57rem">
                  <span class="ok-dot"></span>{info.get('role','—')}
                </span>
              </td>
            </tr>
            <tr><td class="k">Email</td>        <td class="v">{info.get('email','—')}</td></tr>
            <tr><td class="k">Member Since</td> <td class="v">{info.get('joined','—')}</td></tr>
            <tr><td class="k">Session Start</td><td class="v">{sess}</td></tr>
            <tr><td class="k">File Loaded</td>
                <td class="v">{fname if fname else '<span style="color:var(--t3)">None</span>'}</td></tr>
          </table>
        </div>
        """, unsafe_allow_html=True)

        activity = []
        if msgs > 0:
            activity.append(("Sent chat messages this session", f"{msgs} messages"))
        if fname:
            activity.append((f"File uploaded: {fname}", "This session"))
        activity.append(("Session started successfully", sess))
        activity.append(("Authenticated with local system", "Secure login"))

        items_html = ""
        for text, time in activity:
            items_html += f"""
            <div class="act-item">
              <div class="act-dot"></div>
              <div>
                <div class="act-text">{text}</div>
                <div class="act-time">{time}</div>
              </div>
            </div>"""

        st.markdown(f"""
        <div class="gc">
          <div style="font-family:var(--mono);font-size:.54rem;color:var(--t3);
               text-transform:uppercase;letter-spacing:.17em;margin-bottom:.75rem">
            <i class="bi bi-activity"></i>&nbsp; Recent Activity
          </div>
          {items_html}
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.info("🔒 Passwords are SHA-256 hashed. Edit users in `src/backend/auth.py`", icon="ℹ️")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PAGE — MODEL INFO
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def page_model():
    ollama_ok = check_ollama_running()

    _topbar("🤖", "t-model", "Model Info", "LLaMA 3 · Ollama · GGUF · 8B")

    if not ollama_ok:
        _offline_banner()
        return

    minfo = get_model_info()

    entries = [
        ("🏷️", "Model",        minfo.get("name",       MODEL_NAME)),
        ("⚙️", "Parameters",   minfo.get("parameters", "8B")),
        ("📦", "Format",       minfo.get("format",     "gguf").upper()),
        ("🗜️", "Quantisation", minfo.get("quant",      "Q4_0")),
        ("✅", "Status",       minfo.get("status",     "—").upper()),
        ("🌡️", "Temperature",  "0.7"),
    ]

    cols = st.columns(len(entries), gap="small")
    for col, (ico, lbl, val) in zip(cols, entries):
        with col:
            st.markdown(
                f'<div class="metric">'
                f'<span class="metric-ico">{ico}</span>'
                f'<div class="metric-lbl">{lbl}</div>'
                f'<div class="metric-val">{val}</div>'
                f'</div>',
                unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3, gap="large")

    with c1:
        st.markdown("""
        <div class="gc">
          <div style="font-family:var(--mono);font-size:.54rem;color:var(--t3);
               text-transform:uppercase;letter-spacing:.16em;margin-bottom:.7rem">
            <i class="bi bi-terminal"></i>&nbsp; Quick Start
          </div>""", unsafe_allow_html=True)
        st.code("bash setup.sh   # run once\nbash run.sh     # every time", language="bash")
        st.markdown('</div>', unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
        <div class="gc">
          <div style="font-family:var(--mono);font-size:.54rem;color:var(--t3);
               text-transform:uppercase;letter-spacing:.16em;margin-bottom:.8rem">
            <i class="bi bi-sliders"></i>&nbsp; Inference Config
          </div>
          <div class="spec-row"><span class="spec-k">Endpoint</span><span class="spec-v">localhost:11434</span></div>
          <div class="spec-row"><span class="spec-k">Context Window</span><span class="spec-v">4096 tokens</span></div>
          <div class="spec-row"><span class="spec-k">Temperature</span><span class="spec-v">0.7</span></div>
          <div class="spec-row"><span class="spec-k">Top-P</span><span class="spec-v">0.9</span></div>
          <div class="spec-row"><span class="spec-k">Rep. Penalty</span><span class="spec-v">1.1</span></div>
        </div>""", unsafe_allow_html=True)

    with c3:
        st.markdown(f"""
        <div class="gc">
          <div style="font-family:var(--mono);font-size:.54rem;color:var(--t3);
               text-transform:uppercase;letter-spacing:.16em;margin-bottom:.8rem">
            <i class="bi bi-upload"></i>&nbsp; File Q&A Support
          </div>
          <div class="spec-row"><span class="spec-k">PDF</span><span class="spec-v" style="color:#34d399">✓ Supported</span></div>
          <div class="spec-row"><span class="spec-k">TXT / MD / CSV</span><span class="spec-v" style="color:#34d399">✓ Supported</span></div>
          <div class="spec-row"><span class="spec-k">DOCX</span><span class="spec-v" style="color:#34d399">✓ Supported</span></div>
          <div class="spec-row"><span class="spec-k">XLSX / JSON / PY</span><span class="spec-v" style="color:#34d399">✓ Supported</span></div>
          <div class="spec-row"><span class="spec-k">Images</span><span class="spec-v" style="color:#a78bfa">⚡ Metadata only</span></div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
    <div class="gc">
      <div style="font-family:var(--mono);font-size:.54rem;color:var(--t3);
           text-transform:uppercase;letter-spacing:.16em;margin-bottom:.85rem">
        <i class="bi bi-info-circle"></i>&nbsp; About NEXA Chatbot
      </div>
      <div style="font-family:var(--mono);font-size:.78rem;color:var(--t2);line-height:1.8">
        NEXA Chatbot runs <strong style="color:var(--v2)">LLaMA 3 8B</strong> fully locally via Ollama.
        No data is sent to external servers. It supports multi-turn conversations,
        file upload Q&amp;A (PDF, DOCX, TXT, CSV, XLSX, JSON, Python and more), and streaming responses —
        identical to ChatGPT or Claude, but entirely on your machine.
        The model is quantised to <strong style="color:var(--c2)">Q4_0 GGUF</strong>
        for fast CPU/GPU inference with minimal RAM footprint (~5 GB).
        File reading uses multiple fallback libraries so it works across all environments
        including <strong style="color:var(--acc)">Lightning.ai</strong>.
      </div>
    </div>
    """, unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  MAIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    _init_state()
    _inject_styles()

    if not st.session_state.logged_in:
        render_login()
        return

    render_sidebar()

    page = st.session_state.page
    if   page == "chat":  page_chat()
    elif page == "admin": page_admin()
    elif page == "model": page_model()


if __name__ == "__main__":
    main()